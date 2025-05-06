import imaplib
import email
from email.header import decode_header
import openpyxl
from datetime import datetime, timedelta
import calendar
from bs4 import BeautifulSoup
import re
from tqdm import tqdm
import sys

# Cấu hình tài khoản Gmail
EMAIL = 'email@gmail.com'
APP_PASSWORD = '**** **** **** ****'  # Mật khẩu ứng dụng từ Google
IMAP_SERVER = 'imap.gmail.com'
total_in = 0
total_out = 0

# Lấy tháng và năm từ đối số dòng lệnh
if len(sys.argv) < 3:
    print("Vui lòng truyền tháng và năm dưới dạng: python email_analysis.py <tháng> <năm>")
    sys.exit(1)

month = int(sys.argv[1])
year = int(sys.argv[2])

# Tính ngày đầu và cuối tháng
first_day = datetime(year, month, 1)
last_day = first_day.replace(day=calendar.monthrange(year, month)[1])

# Chuyển sang định dạng IMAP
since_date = first_day.strftime("%d-%b-%Y")
before_date = (last_day + timedelta(days=1)).strftime("%d-%b-%Y")

# Kết nối đến Gmail bằng IMAP
mail = imaplib.IMAP4_SSL(IMAP_SERVER)
mail.login(EMAIL, APP_PASSWORD)
mail.select("inbox")

# Tìm tất cả email từ support@timo.vn trong khoảng thời gian tháng trước
search_criteria = f'(FROM "support@timo.vn" SINCE {since_date} BEFORE {before_date})'
status, messages = mail.search(None, search_criteria)

# Tạo workbook Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Subject", "Date", "Amount", "Balance", "Description"])

# Duyệt qua từng email
email_ids = messages[0].split()
print(f"Found {len(email_ids)} emails from {since_date} to {last_day.strftime('%d-%b-%Y')}")
for eid in tqdm(email_ids, desc="Processing emails", unit="email",  colour="green"):
    res, msg_data = mail.fetch(eid, "(RFC822)")
    if res != 'OK':
        continue

    msg = email.message_from_bytes(msg_data[0][1])
    
    # Decode tiêu đề
    subject_raw = msg["Subject"]
    if subject_raw is None:
        continue

    subject_parts = decode_header(subject_raw)
    subject = ""
    for part, enc in subject_parts:
        if isinstance(part, bytes):
            subject += part.decode(enc or "utf-8", errors="ignore")
        else:
            subject += part

    # Bỏ qua nếu tiêu đề không khớp yêu cầu
    if subject.strip() not in ["Debit Transaction Notice", "Credit Transaction Notice"]:
        continue

    date = msg.get("Date")
    

    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_dispo = str(part.get("Content-Disposition"))

            if "attachment" in content_dispo:
                continue  # Bỏ qua file đính kèm

            # Ưu tiên text/plain, nếu không có thì lấy text/html
            try:
                part_body = part.get_payload(decode=True).decode()
            except:
                part_body = ""

            if content_type == "text/plain":
                body = part_body
                break  # Ưu tiên text/plain
            elif content_type == "text/html" and not body:
                body = part_body  # lấy text/html nếu chưa có body nào
    else:
        try:
            body = msg.get_payload(decode=True).decode()
        except:
            body = ""

    # extract nội dung từ HTML
    soup = BeautifulSoup(body, "html.parser")

    paragraphs = soup.find_all("p")

    debit_or_credit_info = ""
    txn_description = ""

    for p in paragraphs:
        text = p.get_text(strip=True)
        if text.startswith("Your Spend Account has been"):
            debit_or_credit_info = text
        elif text.startswith("Transaction Description:"):
            txn_description = text

    # Gộp nội dung bạn muốn lưu
    extracted_body = f"{debit_or_credit_info}\n{txn_description}"
    

    #handling change
    sign = '+' if "credited" in extracted_body else '-'
    # Lấy số tiền giao dịch
    amount_match = re.search(r'has been (?:debited|credited) ([\d,]+) VND', debit_or_credit_info)
    amount = 0
    if amount_match:
        amount = int(amount_match.group(1).replace(",", ""))
        amount = int(f"{sign}{amount}")
    if amount < 0:
        total_out += abs(amount)
    else:
        total_in += amount
    # Lấy số dư
    balance_match = re.search(r'Current Balance is: ([\d,]+)', debit_or_credit_info, re.IGNORECASE)
    balance = int(balance_match.group(1).replace(",", "")) if balance_match else 0

    # print("====== EMAIL ======")
    # print(f"Processing email: {subject}")
    # print(f"Date: {date[:26]}")

    # print(f"Body: {extracted_body}")  
    # print(f"Amount: {amount} VND")
    # print(f"Balance: {balance} VND")
    # print("===================")  
    ws.append([subject,date[:26],amount,balance,txn_description])  

# Lưu file Excel
wb.save(f"timo_transaction_{month}_{year}.xlsx")
mail.logout()
print(f"Đã lưu email vào file: timo_transaction_{month}_{year}.xlsx")
print("====== TỔNG KẾT ======")
print(f"Tổng tiền vào  : {total_in:,} VND")
print(f"Tổng tiền ra   : {total_out:,} VND")
print(f"Số tiền chênh lệch: {total_in - total_out:,} VND")

last_row = ws.max_row
last_balance = ws.cell(row=last_row, column=4).value  # Cột Balance
print(f"Số dư hiện tại  : {last_balance} VND")
print("===================")
